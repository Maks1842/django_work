import os

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, NamedStyle, Side, Border, PatternFill, Alignment, GradientFill
from datetime import date

from ...app_models import *
from ..magic import do_some_magic, answer_in_the_act
from django.template.loader import render_to_string

from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_yasg2.utils import swagger_auto_schema
from drf_yasg2 import openapi
from rest_framework.permissions import IsAdminUser
from django.http import HttpResponse

'''
Добавление всех рейтингов, по проверке, в сводную таблицу Excel.
'''

class ExportRatingsToExcelAPIView(APIView):
    permission_classes = [IsAdminUser]

    @swagger_auto_schema(
        method='get',
        tags=['Рейтинг'],
        operation_description="Сводная таблица рейтингов в Excel",
        manual_parameters=[
            openapi.Parameter('checking', openapi.IN_QUERY, description="Идентификатор проверки",
                              type=openapi.TYPE_INTEGER)
        ])
    @action(detail=False, methods=['get'], )
    def get(self, request):
        checking_id = request.query_params.get('checking')

        if checking_id == None:
            return Response({'error': 'Данные о проверке не предоставлены'})

        ratings_set = Ratings.objects.values().filter(checking_id=checking_id)
        type_organisation_id = ratings_set[0]['type_organisations_id']

        answers_list = []
        for item in ratings_set:
            type_organisation = item['type_organisations_id']
            organisations_id = item['organisations_id']
            queryset = FormsAct.objects.values().get(type_organisations_id=type_organisation)
            form_json = queryset['act_json']

            act_answer = Answers.objects.filter(checking_id=checking_id, type_organisations_id=type_organisation).get(organisations_id=organisations_id).answers_json

            # form_json = FormsAct.objects.get(type_organisations_id=queryset[0].type_organisation).act_json
            query = Question_Values.objects.values()

            comparison = do_some_magic(form_json, act_answer)
            answers = answer_in_the_act(comparison, query)
            answers_list.append(answers)

        result = ''
        match type_organisation_id:
            case (1 | 10 | 6 | 8):
                result = ratings_culture_to_excel(ratings_set, answers_list)
            case (2 | 3):
                result = ratings_healthcare_to_excel(checking_id, answers_list)
            case (4 | 5 | 7 | 9):
                result = ratings_education_to_excel(ratings_set, answers_list)

        file_pointer = open(result, "rb")
        response = HttpResponse(file_pointer, content_type='application/vnd.openxmlformats-officedocument'
                                                           '.spreadsheetml.sheet;')
        response['Content-Disposition'] = f'attachment; filename=download.xlsx'
        response['Content-Transfer-Encoding'] = 'utf-8'
        os.remove(result)

        return response


def ratings_culture_to_excel(ratings_set, answers_list):
    style = style_excel()

    book_template = load_workbook(filename='./checklist/templates/template_culture.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Сводный рейтинг"

    number_col = 6
    count = 0
    for data in ratings_set:
        organisation = Organisations.objects.get(pk=data['organisations_id']).organisation_name

        answer = answers_list[count]

        count += 1
        number_col += 1

        try:
            rating_3_1_1 = data['ratings_json']['rating_3_1_1']['value']
            rating_3_1_2 = data['ratings_json']['rating_3_1_2']['value']
            rating_3_1_3 = data['ratings_json']['rating_3_1_3']['value']
        except:
            rating_3_1_1 = ''
            rating_3_1_2 = ''
            rating_3_1_3 = ''

        # # Вставить один столбец перед №___
        # sheet.insert_cols(number_col)

        column = number_col

        # Преобразую индекс столбца в его буквенное обозначение: 1 = "А"
        letter = get_column_letter(column)

        # Высота строки
        sheet.row_dimensions[1].height = 60

        # Ширина столбца
        sheet.column_dimensions[letter].width = 35

        # Добавить данные в ячейку
        row = 1
        sheet.cell(row, column).value = organisation
        sheet.cell(row, column).style = style['style_main']

        row = 2
        sheet.cell(row, column).value = data['ratings_json']['rating_total']['value']
        sheet.cell(row, column).style = style['style_1']

        row = 3
        sheet.cell(row, column).value = data['ratings_json']['rating_1']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 4
        sheet.cell(row, column).value = data['ratings_json']['rating_1_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 5
        sheet.cell(row, column).value = data['ratings_json']['stend_count_yes']['value']
        sheet.cell(row, column).style = style['style_5']

        row = 6
        sheet.cell(row, column).value = answer['1'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 7
        sheet.cell(row, column).value = answer['2'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 8
        sheet.cell(row, column).value = answer['3'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 9
        sheet.cell(row, column).value = answer['5'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 10
        sheet.cell(row, column).value = answer['6'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 11
        sheet.cell(row, column).value = answer['7'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 12
        sheet.cell(row, column).value = answer['8'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 13
        sheet.cell(row, column).value = answer['11'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 14
        sheet.cell(row, column).value = answer['13'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 15
        sheet.cell(row, column).value = data['ratings_json']['web_count_yes']['value']
        sheet.cell(row, column).style = style['style_5']

        row = 16
        sheet.cell(row, column).value = answer['1'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 17
        sheet.cell(row, column).value = answer['2'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 18
        sheet.cell(row, column).value = answer['3'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 19
        sheet.cell(row, column).value = answer['4'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 20
        sheet.cell(row, column).value = answer['5'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 21
        sheet.cell(row, column).value = answer['6'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 22
        sheet.cell(row, column).value = answer['7'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 23
        sheet.cell(row, column).value = answer['8'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 24
        sheet.cell(row, column).value = answer['9'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 25
        sheet.cell(row, column).value = answer['10'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 26
        sheet.cell(row, column).value = answer['11'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 27
        sheet.cell(row, column).value = answer['13'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 28
        sheet.cell(row, column).value = data['ratings_json']['rating_1_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 29
        sheet.cell(row, column).value = answer['14'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 30
        sheet.cell(row, column).value = answer['15'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 31
        sheet.cell(row, column).value = answer['16'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 32
        sheet.cell(row, column).value = answer['17'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 33
        sheet.cell(row, column).value = answer['18'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 34
        sheet.cell(row, column).value = answer['19'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 35
        sheet.cell(row, column).value = data['ratings_json']['rating_1_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 36
        sheet.cell(row, column).value = data['ratings_json']['count_person_1_3_stend']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 37
        sheet.cell(row, column).value = data['ratings_json']['count_person_1_3_web']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 38
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 39
        sheet.cell(row, column).value = data['ratings_json']['rating_2']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 40
        sheet.cell(row, column).value = data['ratings_json']['rating_2_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 41
        sheet.cell(row, column).value = answer['20'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 42
        sheet.cell(row, column).value = answer['21'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 43
        sheet.cell(row, column).value = answer['22'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 44
        sheet.cell(row, column).value = answer['23'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 45
        sheet.cell(row, column).value = answer['24'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 46
        sheet.cell(row, column).value = answer['25'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 47
        sheet.cell(row, column).value = data['ratings_json']['rating_2_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 48
        sheet.cell(row, column).value = data['ratings_json']['count_person_2_3']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 49
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 50
        sheet.cell(row, column).value = data['ratings_json']['rating_3']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 51
        sheet.cell(row, column).value = data['ratings_json']['rating_3_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 52
        sheet.cell(row, column).style = style['style_6']

        if rating_3_1_1:
            row = 53
            sheet.cell(row, column).style = style['style_main']

            row = 54
            sheet.cell(row, column).style = style['style_main']

            row = 55
            sheet.cell(row, column).style = style['style_main']

            row = 56
            sheet.cell(row, column).value = answer['34'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 57
            sheet.cell(row, column).value = answer['35'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 58
            sheet.cell(row, column).style = style['style_6']

            row = 59
            sheet.cell(row, column).value = rating_3_1_1
            sheet.cell(row, column).style = style['style_4']

            row = 60
            sheet.cell(row, column).value = answer['26'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 61
            sheet.cell(row, column).value = answer['27'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 62
            sheet.cell(row, column).value = answer['28'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 63
            sheet.cell(row, column).value = answer['29'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 64
            sheet.cell(row, column).value = rating_3_1_2
            sheet.cell(row, column).style = style['style_4']

            row = 65
            sheet.cell(row, column).value = answer['30'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 66
            sheet.cell(row, column).value = rating_3_1_3
            sheet.cell(row, column).style = style['style_4']

            row = 67
            sheet.cell(row, column).value = answer['31'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 68
            sheet.cell(row, column).value = answer['32'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 69
            sheet.cell(row, column).value = answer['33'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 70
            sheet.cell(row, column).value = data['ratings_json']['rating_3_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 71
            sheet.cell(row, column).value = answer['36'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 72
            sheet.cell(row, column).value = answer['37'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 73
            sheet.cell(row, column).value = answer['38'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 74
            sheet.cell(row, column).value = answer['39'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 75
            sheet.cell(row, column).value = answer['40'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 76
            sheet.cell(row, column).value = answer['41'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 77
            sheet.cell(row, column).value = data['ratings_json']['rating_3_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 78
            sheet.cell(row, column).value = data['ratings_json']['invalid_person']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 79
            sheet.cell(row, column).value = data['ratings_json']['count_invalid_person_3_3']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 80
            sheet.cell(row, column).value = data['ratings_json']['rating_4']['value']
            sheet.cell(row, column).style = style['style_3']

            row = 81
            sheet.cell(row, column).value = data['ratings_json']['rating_4_1']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 82
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 83
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_1']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 84
            sheet.cell(row, column).value = data['ratings_json']['rating_4_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 85
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 86
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_2']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 87
            sheet.cell(row, column).value = data['ratings_json']['rating_4_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 88
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 89
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_3']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 90
            sheet.cell(row, column).value = data['ratings_json']['rating_5']['value']
            sheet.cell(row, column).style = style['style_3']

            row = 91
            sheet.cell(row, column).value = data['ratings_json']['rating_5_1']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 92
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 93
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_1']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 94
            sheet.cell(row, column).value = data['ratings_json']['rating_5_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 95
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 96
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_2']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 97
            sheet.cell(row, column).value = data['ratings_json']['rating_5_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 98
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 99
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_3']['value']
            sheet.cell(row, column).style = style['style_main']

        else:
            row = 53
            sheet.cell(row, column).value = answer['26'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 54
            sheet.cell(row, column).value = answer['27'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 55
            sheet.cell(row, column).value = answer['28'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 56
            sheet.cell(row, column).value = answer['29'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 57
            sheet.cell(row, column).value = answer['30'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 58
            sheet.cell(row, column).style = style['style_6']

            row = 59
            sheet.cell(row, column).value = rating_3_1_1
            sheet.cell(row, column).style = style['style_4']

            row = 60
            sheet.cell(row, column).style = style['style_main']

            row = 61
            sheet.cell(row, column).style = style['style_main']

            row = 62
            sheet.cell(row, column).style = style['style_main']

            row = 63
            sheet.cell(row, column).style = style['style_main']

            row = 64
            sheet.cell(row, column).value = rating_3_1_2
            sheet.cell(row, column).style = style['style_4']

            row = 65
            sheet.cell(row, column).style = style['style_main']

            row = 66
            sheet.cell(row, column).value = rating_3_1_3
            sheet.cell(row, column).style = style['style_4']

            row = 67
            sheet.cell(row, column).style = style['style_main']

            row = 68
            sheet.cell(row, column).style = style['style_main']

            row = 69
            sheet.cell(row, column).style = style['style_main']

            row = 70
            sheet.cell(row, column).value = data['ratings_json']['rating_3_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 71
            sheet.cell(row, column).value = answer['31'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 72
            sheet.cell(row, column).value = answer['32'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 73
            sheet.cell(row, column).value = answer['33'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 74
            sheet.cell(row, column).value = answer['34'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 75
            sheet.cell(row, column).value = answer['35'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 76
            sheet.cell(row, column).value = answer['36'][0]
            sheet.cell(row, column).style = style['style_main']

            row = 77
            sheet.cell(row, column).value = data['ratings_json']['rating_3_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 78
            sheet.cell(row, column).value = data['ratings_json']['invalid_person']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 79
            sheet.cell(row, column).value = data['ratings_json']['count_invalid_person_3_3']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 80
            sheet.cell(row, column).value = data['ratings_json']['rating_4']['value']
            sheet.cell(row, column).style = style['style_3']

            row = 81
            sheet.cell(row, column).value = data['ratings_json']['rating_4_1']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 82
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 83
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_1']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 84
            sheet.cell(row, column).value = data['ratings_json']['rating_4_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 85
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 86
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_2']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 87
            sheet.cell(row, column).value = data['ratings_json']['rating_4_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 88
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 89
            sheet.cell(row, column).value = data['ratings_json']['count_person_4_3']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 90
            sheet.cell(row, column).value = data['ratings_json']['rating_5']['value']
            sheet.cell(row, column).style = style['style_3']

            row = 91
            sheet.cell(row, column).value = data['ratings_json']['rating_5_1']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 92
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 93
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_1']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 94
            sheet.cell(row, column).value = data['ratings_json']['rating_5_2']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 95
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 96
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_2']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 97
            sheet.cell(row, column).value = data['ratings_json']['rating_5_3']['value']
            sheet.cell(row, column).style = style['style_2']

            row = 98
            sheet.cell(row, column).value = data['ratings_json']['quota']['value']
            sheet.cell(row, column).style = style['style_main']

            row = 99
            sheet.cell(row, column).value = data['ratings_json']['count_person_5_3']['value']
            sheet.cell(row, column).style = style['style_main']

    current_date = date.today()
    file = f'./checklist/local_storage/totalrating_{current_date.strftime("%d.%m.%Y")}.xlsx'
    book_template.save(file)

    return file


def ratings_education_to_excel(ratings_set, answers_list):
    style = style_excel()

    book_template = load_workbook(filename='./checklist/templates/template_education.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Сводный рейтинг"

    number_col = 6
    count = 0
    for data in ratings_set:
        organisation = Organisations.objects.get(pk=data['organisations_id']).organisation_name

        answer = answers_list[count]

        type_organisation_id = data['type_organisations_id']
        type_organisation = Type_Organisations.objects.get(pk=type_organisation_id).type

        count += 1
        number_col += 1

        # # Вставить один столбец перед №___
        # sheet.insert_cols(number_col)

        column = number_col

        # Преобразую индекс столбца в его буквенное обозначение: 1 = "А"
        letter = get_column_letter(column)

        # Высота строки
        sheet.row_dimensions[1].height = 60

        # Ширина столбца
        sheet.column_dimensions[letter].width = 35

        # Добавить данные в ячейку
        row = 1
        sheet.cell(row, column).value = type_organisation
        sheet.cell(row, column).style = style['style_main']

        row = 2
        sheet.cell(row, column).value = organisation
        sheet.cell(row, column).style = style['style_main']

        row = 3
        sheet.cell(row, column).value = data['ratings_json']['rating_total']['value']
        sheet.cell(row, column).style = style['style_1']

        row = 4
        sheet.cell(row, column).value = data['ratings_json']['rating_1']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 5
        sheet.cell(row, column).value = data['ratings_json']['rating_1_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 6
        sheet.cell(row, column).value = data['ratings_json']['stend_count_yes']['value']
        sheet.cell(row, column).style = style['style_5']

        row = 7
        sheet.cell(row, column).value = answer['3'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 8
        sheet.cell(row, column).value = answer['4'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 9
        sheet.cell(row, column).value = answer['5'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 10
        sheet.cell(row, column).value = answer['6'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 11
        sheet.cell(row, column).value = answer['9'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 12
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['10'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 13
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['12'][0]
        else:
            sheet.cell(row, column).value = answer['11'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 14
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['42'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['37'][0]
        else:
            sheet.cell(row, column).value = answer['35'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 15
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['20'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 16
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['22'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['20'][0]
        else:
            sheet.cell(row, column).value = answer['19'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 17
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['29'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 19
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['31'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['27'][0]
        else:
            sheet.cell(row, column).value = answer['26'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 20
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['35'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['31'][0]
        else:
            sheet.cell(row, column).value = answer['30'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 21
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['42'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['37'][0]
        else:
            sheet.cell(row, column).value = answer['35'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 22
        sheet.cell(row, column).value = data['ratings_json']['web_count_yes']['value']
        sheet.cell(row, column).style = style['style_5']

        row = 23
        sheet.cell(row, column).value = answer['1'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 24
        sheet.cell(row, column).value = answer['2'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 25
        sheet.cell(row, column).value = answer['3'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 26
        sheet.cell(row, column).value = answer['4'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 27
        sheet.cell(row, column).value = answer['5'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 28
        sheet.cell(row, column).value = answer['6'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 29
        sheet.cell(row, column).value = answer['7'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 30
        sheet.cell(row, column).value = answer['8'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 31
        sheet.cell(row, column).value = answer['9'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 32
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['10'][1]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 33
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['11'][0]
        else:
            sheet.cell(row, column).value = answer['10'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 34
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['12'][1]
        else:
            sheet.cell(row, column).value = answer['11'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 35
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['13'][0]
        else:
            sheet.cell(row, column).value = answer['12'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 36
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['42'][1]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['37'][1]
        else:
            sheet.cell(row, column).value = answer['35'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 37
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['15'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['14'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 38
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['16'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['15'][0]
        else:
            sheet.cell(row, column).value = answer['14'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 39
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['17'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['16'][0]
        else:
            sheet.cell(row, column).value = answer['15'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 40
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['18'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['17'][0]
        else:
            sheet.cell(row, column).value = answer['16'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 41
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['19'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['18'][0]
        else:
            sheet.cell(row, column).value = answer['17'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 42
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['20'][1]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 43
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['21'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['19'][0]
        else:
            sheet.cell(row, column).value = answer['18'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 44
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['22'][1]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['20'][1]
        else:
            sheet.cell(row, column).value = answer['19'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 45
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['23'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['21'][0]
        else:
            sheet.cell(row, column).value = answer['20'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 46
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['24'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['22'][0]
        else:
            sheet.cell(row, column).value = answer['21'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 47
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['25'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['23'][0]
        else:
            sheet.cell(row, column).value = answer['22'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 48
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['26'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['24'][0]
        else:
            sheet.cell(row, column).value = answer['23'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 49
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['27'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 50
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['28'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['25'][0]
        else:
            sheet.cell(row, column).value = answer['24'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 51
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['29'][1]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 56
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['30'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['26'][0]
        else:
            sheet.cell(row, column).value = answer['25'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 57
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['31'][1]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['27'][1]
        else:
            sheet.cell(row, column).value = answer['26'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 58
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['32'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['28'][0]
        else:
            sheet.cell(row, column).value = answer['27'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 59
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['33'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['29'][0]
        else:
            sheet.cell(row, column).value = answer['28'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 60
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['34'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['30'][0]
        else:
            sheet.cell(row, column).value = answer['29'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 61
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['35'][1]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['31'][1]
        else:
            sheet.cell(row, column).value = answer['30'][1]
        sheet.cell(row, column).style = style['style_main']

        row = 62
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['36'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['32'][0]
        else:
            sheet.cell(row, column).value = answer['31'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 63
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['37'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['33'][0]
        else:
            sheet.cell(row, column).value = answer['32'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 64
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['38'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['34'][0]
        else:
            sheet.cell(row, column).value = answer['33'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 65
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['39'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['35'][0]
        else:
            sheet.cell(row, column).value = answer['34'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 66
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['40'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 67
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['41'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['36'][0]
        else:
            sheet.cell(row, column).value = 'Не предусмотрено'
        sheet.cell(row, column).style = style['style_main']

        row = 70
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['43'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['38'][0]
        else:
            sheet.cell(row, column).value = answer['36'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 71
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['44'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['39'][0]
        else:
            sheet.cell(row, column).value = answer['37'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 72
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['45'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['40'][0]
        else:
            sheet.cell(row, column).value = answer['38'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 73
        sheet.cell(row, column).value = data['ratings_json']['rating_1_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 74
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['46'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['41'][0]
        else:
            sheet.cell(row, column).value = answer['39'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 75
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['47'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['42'][0]
        else:
            sheet.cell(row, column).value = answer['40'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 76
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['48'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['43'][0]
        else:
            sheet.cell(row, column).value = answer['41'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 77
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['49'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['44'][0]
        else:
            sheet.cell(row, column).value = answer['42'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 78
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['50'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['45'][0]
        else:
            sheet.cell(row, column).value = answer['43'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 79
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['51'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['46'][0]
        else:
            sheet.cell(row, column).value = answer['44'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 80
        sheet.cell(row, column).value = data['ratings_json']['rating_1_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 81
        sheet.cell(row, column).value = data['ratings_json']['count_person_1_3_stend']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 82
        sheet.cell(row, column).value = data['ratings_json']['count_person_1_3_web']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 83
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 84
        sheet.cell(row, column).value = data['ratings_json']['rating_2']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 85
        sheet.cell(row, column).value = data['ratings_json']['rating_2_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 86
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['52'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['47'][0]
        else:
            sheet.cell(row, column).value = answer['45'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 87
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['53'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['48'][0]
        else:
            sheet.cell(row, column).value = answer['46'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 88
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['54'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['49'][0]
        else:
            sheet.cell(row, column).value = answer['47'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 89
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['55'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['50'][0]
        else:
            sheet.cell(row, column).value = answer['48'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 90
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['56'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['51'][0]
        else:
            sheet.cell(row, column).value = answer['49'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 91
        sheet.cell(row, column).value = data['ratings_json']['rating_2_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 92
        sheet.cell(row, column).value = data['ratings_json']['count_person_2_3']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 93
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 94
        sheet.cell(row, column).value = data['ratings_json']['rating_3']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 95
        sheet.cell(row, column).value = data['ratings_json']['rating_3_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 96
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['57'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['52'][0]
        else:
            sheet.cell(row, column).value = answer['50'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 97
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['58'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['53'][0]
        else:
            sheet.cell(row, column).value = answer['51'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 98
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['59'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['54'][0]
        else:
            sheet.cell(row, column).value = answer['52'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 99
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['60'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['55'][0]
        else:
            sheet.cell(row, column).value = answer['53'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 100
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['61'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['56'][0]
        else:
            sheet.cell(row, column).value = answer['54'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 101
        sheet.cell(row, column).value = data['ratings_json']['rating_3_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 102
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['62'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['57'][0]
        else:
            sheet.cell(row, column).value = answer['55'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 103
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['63'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['58'][0]
        else:
            sheet.cell(row, column).value = answer['56'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 104
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['64'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['59'][0]
        else:
            sheet.cell(row, column).value = answer['57'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 105
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['65'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['60'][0]
        else:
            sheet.cell(row, column).value = answer['58'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 106
        if type_organisation_id == 5:
            sheet.cell(row, column).value = answer['66'][0]
        elif type_organisation_id == 4:
            sheet.cell(row, column).value = answer['61'][0]
        else:
            sheet.cell(row, column).value = answer['59'][0]
        sheet.cell(row, column).style = style['style_main']

        row = 107
        sheet.cell(row, column).value = data['ratings_json']['rating_3_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 108
        sheet.cell(row, column).value = data['ratings_json']['invalid_person']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 109
        sheet.cell(row, column).value = data['ratings_json']['count_invalid_person_3_3']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 110
        sheet.cell(row, column).value = data['ratings_json']['rating_4']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 111
        sheet.cell(row, column).value = data['ratings_json']['rating_4_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 112
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 113
        sheet.cell(row, column).value = data['ratings_json']['count_person_4_1']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 114
        sheet.cell(row, column).value = data['ratings_json']['rating_4_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 115
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 116
        sheet.cell(row, column).value = data['ratings_json']['count_person_4_2']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 117
        sheet.cell(row, column).value = data['ratings_json']['rating_4_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 118
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 119
        sheet.cell(row, column).value = data['ratings_json']['count_person_4_3']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 120
        sheet.cell(row, column).value = data['ratings_json']['rating_5']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 121
        sheet.cell(row, column).value = data['ratings_json']['rating_5_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 122
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 123
        sheet.cell(row, column).value = data['ratings_json']['count_person_5_1']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 124
        sheet.cell(row, column).value = data['ratings_json']['rating_5_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 125
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 126
        sheet.cell(row, column).value = data['ratings_json']['count_person_5_2']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 127
        sheet.cell(row, column).value = data['ratings_json']['rating_5_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 128
        sheet.cell(row, column).value = data['ratings_json']['quota']['value']
        sheet.cell(row, column).style = style['style_main']

        row = 129
        sheet.cell(row, column).value = data['ratings_json']['count_person_5_3']['value']
        sheet.cell(row, column).style = style['style_main']

    current_date = date.today()
    file = f'./checklist/local_storage/totalrating_{current_date.strftime("%d.%m.%Y")}.xlsx'
    book_template.save(file)

    return file


def ratings_healthcare_to_excel(checking_id, answers_list):

    ratings_set = Ratings.objects.values().filter(checking_id=checking_id)

    organisation_id_list = []
    for item in ratings_set:
        organisation_id_list.append(item['organisations_id'])

    organisation_unique_id = set(organisation_id_list)

    ratings_list = []
    for org_id in organisation_unique_id:
        ratings_org = Ratings.objects.values().filter(checking_id=checking_id, organisations_id=org_id)

        if len(ratings_org) > 1:
            ratings_consolidat = calculating_ratings_consolidat(ratings_org)
            ratings_list.append(ratings_consolidat)
            for r in ratings_org:
                ratings_list.append(r)
        else:
            ratings_list.append(ratings_org[0])

    style = style_excel()

    book_template = load_workbook(filename='./checklist/templates/template_healthcare.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Сводный рейтинг"

    number_col = 1
    for data in ratings_list:
        organisation = Organisations.objects.get(pk=data['organisations_id']).organisation_name

        try:
            # type_organisation_id = Form_Type_Organisation.objects.get(organisation=data['organisations_id']).type_organisation_id
            type_organisation = Type_Organisations.objects.get(pk=data['type_organisations_id']).type
        except:
            type_organisation = 'Консолидированный по организации'

        number_col += 1

        column = number_col

        # Преобразую индекс столбца в его буквенное обозначение: 1 = "А"
        letter = get_column_letter(column)

        # Высота строки
        sheet.row_dimensions[1].height = 60

        # Ширина столбца
        sheet.column_dimensions[letter].width = 35

        # Добавить данные в ячейку
        row = 1
        sheet.cell(row, column).value = type_organisation
        sheet.cell(row, column).style = style['style_main']

        row = 2
        sheet.cell(row, column).value = organisation
        sheet.cell(row, column).style = style['style_main']

        row = 3
        sheet.cell(row, column).value = data['ratings_json']['rating_total']['value']
        sheet.cell(row, column).style = style['style_1']

        row = 4
        sheet.cell(row, column).value = data['ratings_json']['rating_1']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 5
        sheet.cell(row, column).value = data['ratings_json']['rating_1_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 6
        sheet.cell(row, column).value = data['ratings_json']['rating_1_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 7
        sheet.cell(row, column).value = data['ratings_json']['rating_1_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 8
        sheet.cell(row, column).value = data['ratings_json']['rating_2']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 9
        sheet.cell(row, column).value = data['ratings_json']['rating_2_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 10
        sheet.cell(row, column).value = data['ratings_json']['rating_2_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 11
        sheet.cell(row, column).value = data['ratings_json']['rating_2_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 12
        sheet.cell(row, column).value = data['ratings_json']['rating_3']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 13
        sheet.cell(row, column).value = data['ratings_json']['rating_3_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 14
        sheet.cell(row, column).value = data['ratings_json']['rating_3_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 15
        sheet.cell(row, column).value = data['ratings_json']['rating_3_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 16
        sheet.cell(row, column).value = data['ratings_json']['rating_4']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 17
        sheet.cell(row, column).value = data['ratings_json']['rating_4_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 18
        sheet.cell(row, column).value = data['ratings_json']['rating_4_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 19
        sheet.cell(row, column).value = data['ratings_json']['rating_4_3']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 20
        sheet.cell(row, column).value = data['ratings_json']['rating_5']['value']
        sheet.cell(row, column).style = style['style_3']

        row = 21
        sheet.cell(row, column).value = data['ratings_json']['rating_5_1']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 22
        sheet.cell(row, column).value = data['ratings_json']['rating_5_2']['value']
        sheet.cell(row, column).style = style['style_2']

        row = 23
        sheet.cell(row, column).value = data['ratings_json']['rating_5_3']['value']
        sheet.cell(row, column).style = style['style_2']

    current_date = date.today()
    file = f'./checklist/local_storage/totalrating_{current_date.strftime("%d.%m.%Y")}.xlsx'
    book_template.save(file)

    return file


def calculating_ratings_consolidat(ratings_org):

    ratings_0 = ratings_org[0]['ratings_json']
    ratings_1 = ratings_org[1]['ratings_json']

    coefficient = Coefficients.objects.values().get(type_departments=1)

    cfcnt_main = coefficient['main_json']

    quota = ratings_0['quota']['value'] + ratings_1['quota']['value']
    invalid_person = ratings_0['invalid_person']['value'] + ratings_1['invalid_person']['value']

    # Нормативное количество документов на Стенде и Сайте
    stend_count_all = ratings_0['stend_count_all']['value']
    web_count_all = ratings_0['web_count_all']['value']

    stend_count_yes = (ratings_0['stend_count_yes']['value'] + ratings_1['stend_count_yes']['value'])/2
    web_count_yes = (ratings_0['web_count_yes']['value'] + ratings_1['web_count_yes']['value'])/2

    count_person_1_3_stend = ratings_0['count_person_1_3_stend']['value'] + ratings_1['count_person_1_3_stend']['value']
    count_person_1_3_web = ratings_0['count_person_1_3_web']['value'] + ratings_1['count_person_1_3_web']['value']

    rating_1_1 = round((stend_count_yes/stend_count_all + web_count_yes/web_count_all)/2 * 100, 0)

    rating_1_2 = ratings_0['rating_1_2']['value']

    rating_1_3 = round((count_person_1_3_stend + count_person_1_3_web)/(quota * 2) * 100, 0)

    rating_2_1 = (ratings_0['rating_2_1']['value'] + ratings_1['rating_2_1']['value'])/2

    rating_2_2_1 = (ratings_0['rating_2_2_1']['value'] + ratings_1['rating_2_2_1']['value'])/2

    count_person_2_2_2 = ratings_0['count_person_2_2_2']['value'] + ratings_1['count_person_2_2_2']['value']
    rating_2_2_2 = round(count_person_2_2_2/quota * 100, 0)

    count_person_2_3 = ratings_0['count_person_2_3']['value'] + ratings_1['count_person_2_3']['value']
    rating_2_3 = round(count_person_2_3/quota * 100, 0)

    rating_3_1 = (ratings_0['rating_3_1']['value'] + ratings_1['rating_3_1']['value'])/2

    rating_3_2 = (ratings_0['rating_3_2']['value'] + ratings_1['rating_3_2']['value'])/2

    count_invalid_person_3_3 = ratings_0['count_invalid_person_3_3']['value'] + ratings_1['count_invalid_person_3_3']['value']
    rating_3_3 = round(count_invalid_person_3_3/invalid_person * 100, 0)

    count_person_4_1 = ratings_0['count_person_4_1']['value'] + ratings_1['count_person_4_1']['value']
    rating_4_1 = round(count_person_4_1/quota * 100, 0)

    count_person_4_2 = ratings_0['count_person_4_2']['value'] + ratings_1['count_person_4_2']['value']
    rating_4_2 = round(count_person_4_2/quota * 100, 0)

    count_person_4_3 = ratings_0['count_person_4_3']['value'] + ratings_1['count_person_4_3']['value']
    rating_4_3 = round(count_person_4_3/quota * 100, 0)

    count_person_5_1 = ratings_0['count_person_5_1']['value'] + ratings_1['count_person_5_1']['value']
    rating_5_1 = round(count_person_5_1/quota * 100, 0)

    count_person_5_2 = ratings_0['count_person_5_2']['value'] + ratings_1['count_person_5_2']['value']
    rating_5_2 = round(count_person_5_2/quota * 100, 0)

    count_person_5_3 = ratings_0['count_person_5_3']['value'] + ratings_1['count_person_5_3']['value']
    rating_5_3 = round(count_person_5_3/quota * 100, 0)

    rating_1 = round(rating_1_1 * float(cfcnt_main["k_1_1"]) + rating_1_2 * float(cfcnt_main["k_1_2"]) + rating_1_3 * float(cfcnt_main["k_1_3"]), 1)
    rating_2_2 = round(rating_2_2_1 * float(cfcnt_main["k_2_2_1"]) + rating_2_2_2 * float(cfcnt_main["k_2_2_2"]), 0)
    rating_2 = round(rating_2_1 * float(cfcnt_main["k_2_1"]) + rating_2_2 * float(cfcnt_main["k_2_2"]) + rating_2_3 * float(cfcnt_main["k_2_3"]), 1)
    rating_3 = round(rating_3_1 * float(cfcnt_main["k_3_1"]) + rating_3_2 * float(cfcnt_main["k_3_2"]) + rating_3_3 * float(cfcnt_main["k_3_3"]), 1)
    rating_4 = round(rating_4_1 * float(cfcnt_main["k_4_1"]) + rating_4_2 * float(cfcnt_main["k_4_2"]) + rating_4_3 * float(cfcnt_main["k_4_3"]), 1)
    rating_5 = round(rating_5_1 * float(cfcnt_main["k_5_1"]) + rating_5_2 * float(cfcnt_main["k_5_2"]) + rating_5_3 * float(cfcnt_main["k_5_3"]), 1)

    rating_total = round((rating_1 + rating_2 + rating_3 + rating_4 + rating_5)/5, 2)

    ratings_consolidate = {
        "quota": {"value": quota},
        "invalid_person": {"value": invalid_person},
        "rating_total": {"value": rating_total},
        "rating_1": {"value": rating_1},
        "rating_2": {"value": rating_2},
        "rating_3": {"value": rating_3},
        "rating_4": {"value": rating_4},
        "rating_5": {"value": rating_5},

        "rating_1_1": {"value": rating_1_1},
        "stend_count_yes": {"value": stend_count_yes},
        "stend_count_all": {"value": stend_count_all},
        "web_count_yes": {"value": web_count_yes},
        "web_count_all": {"value": web_count_all},
        "rating_1_2": {"value": rating_1_2},
        "rating_1_3": {"value": rating_1_3},
        "count_person_1_3_stend": {"value": count_person_1_3_stend},
        "count_person_1_3_web": {"value": count_person_1_3_web},

        "rating_2_1": {"value": rating_2_1},
        "rating_2_2": {"value": rating_2_2},
        "rating_2_2_1": {"value": rating_2_2_1},
        "rating_2_2_2": {"value": rating_2_2_2},
        "count_person_2_2_2": {"value": count_person_2_2_2},
        "rating_2_3": {"value": rating_2_3},
        "count_person_2_3": {"value": count_person_2_3},

        "rating_3_1": {"value": rating_3_1},
        "rating_3_2": {"value": rating_3_2},
        "rating_3_3": {"value": rating_3_3},
        "count_invalid_person_3_3": {"value": count_invalid_person_3_3},

        "rating_4_1": {"value": rating_4_1},
        "count_person_4_1": {"value": count_person_4_1},
        "rating_4_2": {"value": rating_4_2},
        "count_person_4_2": {"value": count_person_4_2},
        "rating_4_3": {"value": rating_4_3},
        "count_person_4_3": {"value": count_person_4_3},

        "rating_5_1": {"value": rating_5_1},
        "count_person_5_1": {"value": count_person_5_1},
        "rating_5_2": {"value": rating_5_2},
        "count_person_5_2": {"value": count_person_5_2},
        "rating_5_3": {"value": rating_5_3},
        "count_person_5_3": {"value": count_person_5_3},
    }

    return {'organisations_id': ratings_org[0]['organisations_id'], 'ratings_json': ratings_consolidate}


def style_excel():
    # Стиль данных и ячеек
    style_main = NamedStyle(name="style_main")
    style_main.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_main.border = Border(left=side, right=side, top=side, bottom=side)
    style_main.fill = PatternFill("solid", fgColor="FFFFFF")
    style_main.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_1 = NamedStyle(name="style_1")
    style_1.font = Font(bold=True, size=12, color="FFFFFF")
    side = Side(style='thin', color="000000")
    style_1.border = Border(left=side, right=side, top=side, bottom=side)
    style_1.fill = PatternFill("solid", fgColor="2F5597")
    style_1.alignment = Alignment(horizontal='center', vertical='center')

    style_2 = NamedStyle(name="style_2")
    style_2.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_2.border = Border(left=side, right=side, top=side, bottom=side)
    style_2.fill = PatternFill("solid", fgColor="c5e0b4")
    style_2.alignment = Alignment(horizontal='center', vertical='center')

    style_3 = NamedStyle(name="style_3")
    style_3.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_3.border = Border(left=side, right=side, top=side, bottom=side)
    style_3.fill = PatternFill("solid", fgColor="b4c7e7")
    style_3.alignment = Alignment(horizontal='center', vertical='center')

    style_4 = NamedStyle(name="style_4")
    style_4.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_4.border = Border(left=side, right=side, top=side, bottom=side)
    style_4.fill = PatternFill("solid", fgColor="afd095")
    style_4.alignment = Alignment(horizontal='center', vertical='center')

    style_5 = NamedStyle(name="style_5")
    style_5.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_5.border = Border(left=side, right=side, top=side, bottom=side)
    style_5.fill = PatternFill("solid", fgColor="ffe699")
    style_5.alignment = Alignment(horizontal='center', vertical='center')

    style_6 = NamedStyle(name="style_6")
    style_6.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_6.border = Border(left=side, right=side, top=side, bottom=side)
    style_6.fill = PatternFill("solid", fgColor="f8cbad")
    style_6.alignment = Alignment(horizontal='center', vertical='center')

    return {
        'style_main': style_main,
        'style_1': style_1,
        'style_2': style_2,
        'style_3': style_3,
        'style_4': style_4,
        'style_5': style_5,
        'style_6': style_6,
    }



