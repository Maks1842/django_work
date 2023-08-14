import openpyxl
from openpyxl.styles import Font, NamedStyle, Side, Border, PatternFill, Alignment
from dadata import Dadata
import time
import os
import pandas as pd
import json

''' Модуль получения данных о компании по ИНН, с https://dadata.ru/api/find-party/ по API'''

token = "1fd6b2895fd9a922bf0548395fb09aa1eb4565af"
dadata = Dadata(token)
def data_extract():

    path_file = '/home/maks/Загрузки/inn_obr.xlsx'

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    list_inn = []
    for row in range(1, sheet.max_row):
        list_inn.append(str(sheet[row][0].value))

    # list_inn = ['2634026041', '2634026059', '2636015197']

    result = []
    count = 0
    for inn in list_inn:
        result.append(dadata.find_by_id("party", f"{inn}"))
        count += 1
        time.sleep(0.05)

    # print(result)

    style = style_excel()
    book = openpyxl.Workbook()

    # Первый (активный) Лист книги
    sheet = book.active

    # Изменить имя Листа книги
    sheet.title = "Реестр"

    # Заголовки
    sheet["A1"].value = "Наименование сокращенное"
    sheet["B1"].value = "Наименование полное"
    sheet["C1"].value = "ИНН"
    sheet["D1"].value = "ОГРН"
    sheet["E1"].value = "ОКАТО"
    sheet['F1'].value = "Адрес"
    sheet['G1'].value = "Широта"
    sheet['H1'].value = "Долгота"
    sheet['I1'].value = "Директор"

    # Стили заголовков
    sheet['A1'].style = style['style_1']
    sheet['B1'].style = style['style_1']
    sheet['C1'].style = style['style_1']
    sheet['D1'].style = style['style_1']
    sheet['E1'].style = style['style_1']
    sheet['F1'].style = style['style_1']
    sheet['G1'].style = style['style_1']
    sheet['H1'].style = style['style_1']
    sheet['I1'].style = style['style_1']

    number_col = 1
    number_row = 2
    for item in result:

        # print(item[0]['data']['management']['name'])

        number_col += 1

        # # Вставить один столбец перед №___
        # sheet.insert_cols(number_col)

        column = 1
        row = number_row

        # Высота строки
        sheet.row_dimensions[row].height = 20

        # Ширина столбца
        sheet.column_dimensions["A"].width = 80
        sheet.column_dimensions["B"].width = 100
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
        sheet.column_dimensions["F"].width = 80
        sheet.column_dimensions["G"].width = 20
        sheet.column_dimensions["H"].width = 20
        sheet.column_dimensions["I"].width = 20

        # Добавить данные в ячейку
        column = 1
        sheet.cell(row, column).value = item[0]['value']
        sheet.cell(row, column).style = style['style_main']

        column = 2
        sheet.cell(row, column).value = item[0]['data']['name']['full_with_opf']
        sheet.cell(row, column).style = style['style_main']

        column = 3
        sheet.cell(row, column).value = item[0]['data']['inn']
        sheet.cell(row, column).style = style['style_main']

        column = 4
        sheet.cell(row, column).value = item[0]['data']['ogrn']
        sheet.cell(row, column).style = style['style_main']

        column = 5
        okato = item[0]['data']['okato'][:5] + '000'
        sheet.cell(row, column).value = okato
        sheet.cell(row, column).style = style['style_main']

        column = 6
        sheet.cell(row, column).value = item[0]['data']['address']['unrestricted_value']
        sheet.cell(row, column).style = style['style_main']

        column = 7
        sheet.cell(row, column).value = item[0]['data']['address']['data']['geo_lat']
        sheet.cell(row, column).style = style['style_main']

        column = 8
        sheet.cell(row, column).value = item[0]['data']['address']['data']['geo_lon']
        sheet.cell(row, column).style = style['style_main']

        column = 9
        sheet.cell(row, column).value = item[0]['data']['management']['name']
        sheet.cell(row, column).style = style['style_main']

        number_row += 1

    file = os.path.join('/home/maks/Загрузки', 'Реестр учреждений образования.xlsx')
    book.save(file)

def style_excel():
    # Стиль данных и ячеек
    style_main = NamedStyle(name="style_main")
    style_main.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_main.border = Border(left=side, right=side, top=side, bottom=side)
    style_main.fill = PatternFill("solid", fgColor="FFFFFF")
    style_main.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_1 = NamedStyle(name="style_1")
    style_1.font = Font(bold=True, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_1.border = Border(left=side, right=side, top=side, bottom=side)
    style_1.fill = PatternFill("solid", fgColor="FFFFFF")
    style_1.alignment = Alignment(horizontal='center', vertical='center')

    return {
        'style_main': style_main,
        'style_1': style_1,
    }

# data_extract()


"""
Заголовки столбцов должны быть на латинице, так как преобразуется в json
"""
def import_registry_organisations():

    excel_data = pd.read_excel('/home/maks/Загрузки/Реестр учреждений образования.xlsx')
    json_str = excel_data.to_json(orient='records', date_format='iso')
    parsed = json.loads(json_str)

    print(parsed)

    # return parsed
# import_registry_organisations()
